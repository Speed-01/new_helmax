# ────────────────────────────────────────────────
# Standard library imports
# ────────────────────────────────────────────────
import json
import logging
from datetime import datetime, timedelta
from decimal import Decimal

# ────────────────────────────────────────────────
# Django core imports
# ────────────────────────────────────────────────
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST, require_http_methods
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction, models
from django.db.models import Q, Count, Sum, Min
from django.utils import timezone

# ────────────────────────────────────────────────
# Third-party library imports
# ────────────────────────────────────────────────
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ────────────────────────────────────────────────
# Local application imports
# ────────────────────────────────────────────────
from .models import (
    Product, Category, ProductImage, User, Brand, Variant, Size,
    Coupon, CouponUsage, Order, ReturnRequest, Wallet, WalletTransaction,
    ProductOffer, CategoryOffer, OrderItem, OrderStatusHistory
)
from .pdf_generator import generate_sales_report_pdf
from store.utils import send_order_status_notification

# ────────────────────────────────────────────────
# Project-level imports
# ────────────────────────────────────────────────

from django.conf import settings

# Configure logging
logger = logging.getLogger(__name__)


def search_products(request):
    query = request.GET.get('q', '').strip()
    products = []

    if query:
        # Filter products by name or description
        products = Product.objects.filter(name__icontains=query).values('id', 'name', 'description')

    return JsonResponse({'products': list(products)})



@login_required(login_url='adminLogin')
def get_top_categories(request):
    top_categories = OrderItem.objects.filter(
        order__order_status='DELIVERED'
    ).values(
        'product__category__name'
    ).annotate(
        total_sales=Sum('total_price'),
        total_quantity=Sum('quantity')
    ).order_by('-total_sales')[:5]
    
    return JsonResponse({'categories': list(top_categories)})

@login_required(login_url='adminLogin')
def get_top_products(request):
    top_products = OrderItem.objects.filter(
        order__order_status='DELIVERED'
    ).values(
        'product__name'
    ).annotate(
        total_sales=Sum('total_price'),
        total_quantity=Sum('quantity')
    ).order_by('-total_sales')[:5]
    
    return JsonResponse({'products': list(top_products)})

@login_required(login_url='adminLogin')
def get_top_brands(request):
    top_brands = OrderItem.objects.filter(
        order__order_status='DELIVERED'
    ).values(
        'product__brand__name'
    ).annotate(
        total_sales=Sum('total_price'),
        total_quantity=Sum('quantity')
    ).order_by('-total_sales')[:5]
    
    return JsonResponse({'brands': list(top_brands)})



@never_cache
def adminLogin(request):
    if request.user.is_authenticated:
        return redirect("admin_dashboard")
    
    username_value = ''
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        username_value = username  # Store for re-rendering
        
        user = authenticate(username=username, password=password)
        if user is None:
            messages.error(request, "Invalid username or password")
            return render(request, "adminLogin.html", {'username': username_value})
        elif user and user.is_superuser:
            login(request, user)
            return redirect("admin_dashboard")
        else:
            messages.error(request, f"{user.username} has no access to this page")
            return render(request, "adminLogin.html", {'username': username_value})
    return render(request, "adminLogin.html", {'username': username_value})


@csrf_exempt  
@never_cache  
def admin_logout(request):
    if request.method == 'POST':
        request.session.flush()  # Completely clear session
        logout(request)  # Django logout
        response = redirect('adminLogin')
        response.delete_cookie('sessionid')  # Remove session cookie
        response.delete_cookie('csrftoken')  # Remove CSRF token
        return response
    return redirect('adminLogin')


@never_cache
@login_required(login_url='adminLogin')
def customers(request):
    credential = request.GET.get("value", "")
    if credential:
        customers = User.objects.filter(
            Q(username__icontains=credential) | Q(email__icontains=credential)
        ).exclude(is_superuser=True)
    else:
        customers = User.objects.all().exclude(is_superuser=True).order_by('id')
        
    search_query = request.GET.get('search', '')
    if search_query:
        customers = customers.filter(first_name__icontains=search_query)

    paginator = Paginator(customers, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "customers": page_obj,
        "search_query": search_query
    }

    return render(request, "customers.html", context)



@login_required(login_url='adminLogin')
def toggle_user_status(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        if not user.is_superuser:  
            user.is_active = not user.is_active
            user.save()
            return JsonResponse({
                'success': True,
                'status': 'active' if user.is_active else 'blocked'
            })
        
    return JsonResponse({'success': False}, status=400)
        




@login_required(login_url='adminLogin')
def admin_category(request):
    search_query = request.GET.get('search', '')
    
    # Filter categories based on search query
    categories = Category.objects.all().order_by("-id")
    if search_query:
        categories = categories.filter(name__icontains=search_query)
    
    # Pagination
    page_number = request.GET.get('page', 1)
    paginator = Paginator(categories, 10)  # 10 items per page
    page_obj = paginator.get_page(page_number)
    
    context ={
        'categories': page_obj,
        'search_query': search_query
    }
    
    return render(request, 'admincategory.html', context )

@login_required(login_url='adminLogin')
def add_category(request):
    if request.method == "POST":
        name = request.POST.get('category_name')
        if name:  
            if Category.objects.filter(name__iexact=name).exists():
                messages.error(request, "Category with this name already exists.")
                error_message = "Category with this name already exists."
                return render(request, 'admincategory.html', {
                    'categories': Category.objects.all(),
                    'error_message': error_message,
                    'show_add_modal': True,  
                })
            else:
                Category.objects.create(name=name)
                messages.success(request, "Category added successfully!")
        return redirect('admin_category')  
    return redirect('admin_category')


@login_required(login_url='adminLogin')
def edit_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == "POST":
        name = request.POST.get('category_name')
        if name:
            if Category.objects.filter(name__iexact=name).exclude(id=category_id).exists():
                messages.error(request, "Category with this name already exists.")
                error_message = "Category with this name already exists."
                return render(request, 'admincategory.html', {
                    'categories': Category.objects.all(),
                    'error_message': error_message,
                    'edit_category': category,
                    'show_edit_modal': True,
                })
            else:
                category.name = name
                category.save()
                messages.success(request, "Category updated successfully!")
        return redirect('admin_category')
    return render(request, 'edit_category.html', {'category': category})



@login_required(login_url='adminLogin')
def toggle_category_status(request, category_id):
    print(f"Toggling category {category_id}")
    if request.method == 'POST':
        category = get_object_or_404(Category, id=category_id)
        category.is_active = not category.is_active
        category.save()
        print(f"New status: {'active' if category.is_active else 'blocked'}")
        return JsonResponse({
            'success': True,
            'status': 'active' if category.is_active else 'blocked'
        })
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)


# def delete_category(request, category_id):
#     category = get_object_or_404(Category, id=category_id)
#     category.is_deleted = True
#     category.save()
#     print("Category deleted successfully")
#     return redirect('admin_category')


@login_required(login_url='adminLogin')
def sales_report(request):
    # Get report type and date range from request
    report_type = request.GET.get('report_type', 'daily')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    download_format = request.GET.get('download')

    # Initialize the date range based on report type
    today = timezone.now()
    if report_type == 'daily':
        start_date = today.date()
        end_date = start_date
    elif report_type == 'weekly':
        start_date = today.date() - timezone.timedelta(days=today.weekday())
        end_date = start_date + timezone.timedelta(days=6)
    elif report_type == 'monthly':
        start_date = today.date().replace(day=1)
        end_date = (start_date + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(days=1)
    elif report_type == 'custom' and start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start_date = today.date()
        end_date = start_date

    # Filter orders based on date range and delivery status
    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        order_status='DELIVERED'
    ).order_by('-created_at')

    # Calculate summary statistics
    total_orders = orders.count()
    total_sales = sum(order.total_amount for order in orders)
    total_discounts = sum(order.total_discount for order in orders)

    # Handle report downloads
    if download_format:
        if download_format == 'pdf':
            # Generate PDF report with professional formatting
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="helmax_sales_report_{start_date}_{end_date}.pdf"'
            
            # Use our professional PDF generator function
            return generate_sales_report_pdf(
                response=response,
                orders=orders,
                total_orders=total_orders,
                total_sales=total_sales,
                total_discounts=total_discounts,
                start_date=start_date,
                end_date=end_date
            )
            

            
        elif download_format == 'excel':
            # Generate Excel report
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.xlsx"'
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Sales Report'
            
            # Add headers
            headers = ['Order ID', 'Date', 'Customer', 'Items', 'Subtotal', 'Discount', 'Total', 'Status']
            worksheet.append(headers)
            
            # Add data
            for order in orders:
                worksheet.append([
                    order.order_id,
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.user.username,
                    order.order_items.count(),
                    float(order.total_amount),
                    float(order.total_discount),
                    float(order.total_amount),
                    order.order_status
                ])
            
            # Style the worksheet
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            workbook.save(response)
            return response

    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_sales': total_sales,
        'total_discounts': total_discounts,
        'report_type': report_type,
        'start_date': start_date,
        'end_date': end_date
    }
    
    return render(request, 'sales_report.html', context)

@login_required(login_url='adminLogin')
def admin_brand(request):
    # Fetch all brands ordered by ID in descending order
    brands = Brand.objects.all().order_by("-id")
    
    # Search functionality for brands
    search_query = request.GET.get('search', '')
    if search_query:
        
        brands = brands.filter(name__icontains=search_query)
    
    # Paginate brands
    paginator = Paginator(brands, 10)  # Show 10 brands per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Prepare context
    context = {
        'brands': page_obj,  # Pass the paginated brands
        'search_query': search_query,  # Pass the search query
    }
    
    # Render the template with the context
    return render(request, 'adminBrand.html', context)

@login_required(login_url='adminLogin')
def add_brand(request):
    if request.method == "POST":
        name = request.POST.get('brand_name')
        if name and not Brand.objects.filter(name__iexact=name).exists():
            Brand.objects.create(name=name)
        return redirect('admin_brand')
    return redirect('admin_brand')

@login_required(login_url='adminLogin')
def edit_brand(request, brand_id):
    brand = get_object_or_404(Brand, id=brand_id)
    if request.method == "POST":
        
        name = request.POST.get('brand_name')
        if name and not Brand.objects.filter(name__iexact=name).exclude(id=brand_id).exists():
            brand.name = name
            brand.save()
        return redirect('admin_brand')
    return render(request, 'editBrand.html', {'brand': brand})

@login_required(login_url='adminLogin')
def toggle_brand_status(request, brand_id):
    if request.method == 'POST':
        brand = get_object_or_404(Brand, id=brand_id)
        brand.is_active = not brand.is_active
        brand.save()
        return JsonResponse({
            'success': True,
            'status': 'active' if brand.is_active else 'blocked'
        })
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)


@login_required(login_url='adminLogin')
def adminProducts(request):
    # Fetch all products ordered by ID
    products = Product.objects.all().order_by('id')
    
    # Pagination Logic
    paginator = Paginator(products, 10)  # Show 10 products per page
    page_number = request.GET.get('page')  # Get the current page number from the request
    page_obj = paginator.get_page(page_number)  # Get the page object

    context = {
        'page_obj': page_obj,  # Pass the paginated page object to the template
    }
    return render(request, 'adminProducts.html', context)


@login_required(login_url='adminLogin')
def addProducts(request):
    categories = Category.objects.filter(is_active=True)
    brands = Brand.objects.filter(is_active=True)
    
    if request.method == 'POST':
        # Get and sanitize input
        name = request.POST.get('name', '').strip()  # Remove leading/trailing whitespace
        brand_id = request.POST.get('brand')
        category_id = request.POST.get('category')
        description = request.POST.get('description', '').strip()

        # Validate all fields are not empty
        if not name or not brand_id or not category_id or not description:
            messages.error(request, "All fields are required.")
            return render(request, 'addProducts.html', {'categories': categories, 'brands': brands})

        # # Validate product name (no symbols or only spaces)
        # if not name.replace(" ", "").isalnum():
        #     messages.error(request, "Product name cannot contain only symbols or spaces.")
        #     return render(request, 'addProducts.html', {'categories': categories, 'brands': brands})

        # Case-insensitive uniqueness check
        if Product.objects.filter(name__iexact=name).exists():
            messages.error(request, f"A product with the name '{name}' already exists (case-insensitive).")
            return render(request, 'addProducts.html', {'categories': categories, 'brands': brands})

        # Create product if valid
        try:
            category = get_object_or_404(Category, id=category_id)
            brand = get_object_or_404(Brand, id=brand_id)
            
            Product.objects.create(
                name=name,
                brand=brand,
                category=category,
                description=description
            )
            messages.success(request, "Product added successfully!")
            return redirect('adminProducts')
            
        except Exception as e:
            messages.error(request, f"Error creating product: {str(e)}")

    return render(request, 'addProducts.html', {'categories': categories, 'brands': brands})

@login_required(login_url='adminLogin')
def api_brands(request):
    try:
        # Get query parameters with defaults
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 10))
        sort_field = request.GET.get('sort_field', 'id')
        sort_direction = request.GET.get('sort_direction', 'asc')
        
        # Define valid sort fields based on your model fields
        valid_sort_fields = ['id', 'name', 'is_active']
        
        # Validate sort field to prevent error with non-existent fields
        if sort_field not in valid_sort_fields:
            # Log the invalid sort field attempt
            logger.warning(f"Invalid sort field requested: {sort_field}")
            # Default to id if invalid sort field is requested
            sort_field = 'id'
        
        # Special handling for common sort fields that might be requested but don't exist directly
        # For example, if 'created_at' is requested but doesn't exist, we can fall back to 'id'
        if sort_field == 'created_at':
            logger.info("Sorting by 'created_at' requested, falling back to 'id'")
            sort_field = 'id'
        
        # Apply sorting direction
        if sort_direction == 'desc':
            sort_field = f'-{sort_field}'
        
        # Filter brands based on search query if provided
        brands_query = Brand.objects.all()
        if search:
            brands_query = brands_query.filter(name__icontains=search)
        
        # Apply sorting
        brands_query = brands_query.order_by(sort_field)
        
        # Set up pagination
        paginator = Paginator(brands_query, per_page)
        try:
            brands_page = paginator.page(page)
        except (PageNotAnInteger, EmptyPage):
            # Default to first page if page number is invalid
            brands_page = paginator.page(1)
        
        # Prepare data for JSON response
        brands_data = []
        for brand in brands_page:
            # Count products associated with this brand
            product_count = Product.objects.filter(brand=brand).count()
            
            # Create brand data dictionary
            brands_data.append({
                'id': brand.id,
                'name': brand.name,
                'is_active': brand.is_active,
                'product_count': product_count
            })
        
        # Create response with pagination metadata
        response_data = {
            'items': brands_data,
            'page': brands_page.number,
            'total_pages': paginator.num_pages,
            'total': paginator.count,
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        # Log the error for server-side debugging
        logger.error(f"Error in api_brands view: {str(e)}")
        # Return error response to client
        return JsonResponse({
            'error': 'An error occurred while fetching brands.',
            'details': str(e)
        }, status=500)



logger = logging.getLogger(__name__)

@login_required
@require_http_methods(["GET"])
def api_products(request):
    """
    Improved API endpoint for fetching products with better error handling,
    optimized queries, and proper data validation.
    """
    try:
        # Get and validate query parameters
        search = request.GET.get('search', '').strip()
        page = max(1, int(request.GET.get('page', 1)))
        per_page = min(100, max(1, int(request.GET.get('per_page', 10))))
        sort_field = request.GET.get('sort_field', 'name').strip()
        sort_direction = request.GET.get('sort_direction', 'asc').strip()
        status_filter = request.GET.get('status', 'all').strip()
        category_filter = request.GET.get('category', 'all').strip()
        
        # Define valid sort fields to prevent SQL injection
        VALID_SORT_FIELDS = {
            'id': 'id',
            'name': 'name',
            'brand': 'brand__name',
            'category': 'category__name',
            'is_active': 'is_active',
            'created_at': 'created_at',
            'updated_at': 'updated_at'
        }
        
        # Validate and map sort field
        if sort_field not in VALID_SORT_FIELDS:
            logger.warning(f"Invalid sort field requested: {sort_field}, defaulting to 'name'")
            sort_field = 'name'
        
        sort_field_mapped = VALID_SORT_FIELDS[sort_field]
        
        # Apply sorting direction
        if sort_direction not in ['asc', 'desc']:
            sort_direction = 'asc'
            
        if sort_direction == 'desc':
            sort_field_mapped = f'-{sort_field_mapped}'
        
        # Build the query with optimized select_related and prefetch_related
        products_query = Product.objects.select_related(
            'category', 'brand'
        ).prefetch_related(
            'variants',
            'variants__sizes'
        )
        
        # Apply search filter
        if search:
            products_query = products_query.filter(
                Q(name__icontains=search) |
                Q(brand__name__icontains=search) |
                Q(category__name__icontains=search)
            )
        
        # Apply status filter
        if status_filter in ['active', 'blocked']:
            is_active = status_filter == 'active'
            products_query = products_query.filter(is_active=is_active)
        
        # Apply category filter
        if category_filter != 'all':
            try:
                category_id = int(category_filter)
                products_query = products_query.filter(category_id=category_id)
            except (ValueError, TypeError):
                logger.warning(f"Invalid category filter: {category_filter}")
        
        # Apply sorting
        try:
            products_query = products_query.order_by(sort_field_mapped)
        except Exception as e:
            logger.error(f"Error applying sort: {str(e)}")
            products_query = products_query.order_by('name')
        
        # Set up pagination
        paginator = Paginator(products_query, per_page)
        
        try:
            products_page = paginator.page(page)
        except PageNotAnInteger:
            logger.warning(f"Invalid page number: {page}, defaulting to page 1")
            products_page = paginator.page(1)
            page = 1
        except EmptyPage:
            logger.warning(f"Page {page} is out of range, defaulting to last page")
            products_page = paginator.page(paginator.num_pages)
            page = paginator.num_pages
        
        # Prepare optimized data for JSON response
        products_data = []
        
        for product in products_page:
            try:
                # Get category and brand names safely
                category_name = product.category.name if product.category else "N/A"
                brand_name = product.brand.name if product.brand else "N/A"
                
                # Calculate aggregated data from variants (optimized)
                total_stock = 0
                min_price = None
                
                # Use the prefetched variants to avoid additional queries
                for variant in product.variants.all():
                    # Calculate stock from prefetched sizes
                    variant_stock = sum(size.stock for size in variant.sizes.all())
                    total_stock += variant_stock
                    
                    # Track minimum price
                    if variant.price:
                        if min_price is None or variant.price < min_price:
                            min_price = variant.price
                
                # Create product data dictionary
                product_data = {
                    'id': product.id,
                    'name': product.name,
                    'category': category_name,
                    'brand': brand_name,
                    'price': float(min_price) if min_price else 0.0,
                    'stock': total_stock,
                    'is_active': product.is_active,
                    'created_at': product.created_at.isoformat() if hasattr(product, 'created_at') else None,
                }
                
                products_data.append(product_data)
                
            except Exception as e:
                logger.error(f"Error processing product {product.id}: {str(e)}")
                # Continue with other products even if one fails
                continue
        
        # Create response with comprehensive pagination metadata
        response_data = {
            'items': products_data,
            'page': page,
            'per_page': per_page,
            'total_pages': paginator.num_pages,
            'total': paginator.count,
            'has_next': products_page.has_next(),
            'has_previous': products_page.has_previous(),
            'search': search,
            'status_filter': status_filter,
            'category_filter': category_filter,
            'sort_field': sort_field,
            'sort_direction': sort_direction
        }
        
        return JsonResponse(response_data, safe=False)
        
    except ValueError as e:
        logger.error(f"ValueError in api_products: {str(e)}")
        return JsonResponse({
            'error': 'Invalid parameter value',
            'message': str(e)
        }, status=400)
        
    except Exception as e:
        logger.error(f"Unexpected error in api_products: {str(e)}")
        return JsonResponse({
            'error': 'An unexpected error occurred while fetching products',
            'message': str(e) if settings.DEBUG else 'Internal server error'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def toggle_product_status(request, product_id):
    """
    Improved toggle product status view with better error handling
    and validation.
    """
    try:
        # Validate product_id
        if not product_id:
            return JsonResponse({
                'success': False,
                'error': 'Product ID is required'
            }, status=400)
        
        # Get the product
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Product with ID {product_id} not found'
            }, status=404)
        
        # Toggle the status
        product.is_active = not product.is_active
        product.save(update_fields=['is_active'])
        
        logger.info(f"Product {product_id} status toggled to {'active' if product.is_active else 'blocked'}")
        
        return JsonResponse({
            'success': True,
            'message': f'Product status updated to {"active" if product.is_active else "blocked"}',
            'is_active': product.is_active
        })
        
    except Exception as e:
        logger.error(f"Error toggling product status for ID {product_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating product status',
            'message': str(e) if settings.DEBUG else 'Internal server error'
        }, status=500)


# Alternative optimized query using database aggregation (for better performance)
@login_required
@require_http_methods(["GET"])
def api_products_optimized(request):
    """
    Ultra-optimized version using database aggregation for large datasets
    """
    try:
        # Same parameter validation as above...
        search = request.GET.get('search', '').strip()
        page = max(1, int(request.GET.get('page', 1)))
        per_page = min(100, max(1, int(request.GET.get('per_page', 10))))
        
        # Use database aggregation for better performance
        products_query = Product.objects.select_related(
            'category', 'brand'
        ).annotate(
            total_stock=Sum('variants__sizes__stock'),
            min_price=Min('variants__price')
        )
        
        # Apply filters (same as above)...
        if search:
            products_query = products_query.filter(
                Q(name__icontains=search) |
                Q(brand__name__icontains=search) |
                Q(category__name__icontains=search)
            )
        
        # Pagination
        paginator = Paginator(products_query, per_page)
        products_page = paginator.page
    except Exception as e:
        logger.error(f"Error in api_products_optimized: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)
        
@login_required(login_url='adminLogin')
def addVariant(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        color = request.POST.get('color')
        price = request.POST.get('price')
        discount_price = request.POST.get('discount_price')
        sizes = request.POST.getlist('sizes')
        
        # Handle both original and cropped images
        images = request.FILES.getlist('images')
        cropped_indices = request.POST.get('cropped_indices', '').split(',')
        cropped_indices = [int(i) for i in cropped_indices if i.isdigit()]
        
        primary_image_index = int(request.POST.get('primary_image_index', 0))

        try:
            with transaction.atomic():
                # Create variant
                variant = Variant.objects.create(
                    product=product,
                    color=color,
                    price=price,
                    discount_price=discount_price if discount_price else None
                )

                # Create sizes with their respective stocks
                for size in sizes:
                    stock_for_size = request.POST.get(f'stock_{size}', 0)
                    stock_value = int(stock_for_size) if stock_for_size else 0
                    
                    Size.objects.create(
                        product_variant=variant,
                        name=size,
                        stock=stock_value
                    )

                # Handle images - both original and cropped
                for index, image in enumerate(images):
                    # Check if this image was cropped
                    if index in cropped_indices:
                        # Use the cropped version if available
                        cropped_key = f'cropped_image_{index}'
                        if cropped_key in request.FILES:
                            image_to_save = request.FILES[cropped_key]
                        else:
                            image_to_save = image
                    else:
                        image_to_save = image
                    
                    ProductImage.objects.create(
                        variant=variant,
                        image=image_to_save,
                        is_primary=(index == primary_image_index)
                    )

            messages.success(request, 'Variant added successfully.')
        except Exception as e:
            messages.error(request, f'Error adding variant: {str(e)}')

        return redirect('adminProducts')

    context = {
        'product': product,
        'size_choices': Size.SIZE_CHIOCES
    }
    return render(request, 'addVariant.html', context)

@login_required(login_url='adminLogin')
def editProduct(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    categories = Category.objects.filter(is_active=True)
    brands = Brand.objects.filter(is_active=True)
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            brand_id = request.POST.get('brand')
            category_id = request.POST.get('category')
            description = request.POST.get('description')

            product.name = name
            product.brand = get_object_or_404(Brand, id=brand_id)
            product.category = get_object_or_404(Category, id=category_id)
            product.description = description
            product.save()

            return redirect('adminProducts')
        except Exception as e:
            print(f"Error: {e}")
    
    context = {
        'product': product,
        'categories': categories,
        'brands': brands
    }
    return render(request, 'editProduct.html', context)

@login_required(login_url='adminLogin')
def editVariant(request, variant_id):
    variant = get_object_or_404(Variant, id=variant_id)
    product = variant.product
    existing_sizes = variant.sizes.all()
    existing_images = variant.images.all()
    
    if request.method == 'POST':
        color = request.POST.get('color')
        price = request.POST.get('price')
        discount_price = request.POST.get('discount_price')
        new_sizes = request.POST.getlist('sizes')
        new_images = request.FILES.getlist('images')
        
        # Handle both original and cropped images
        cropped_indices = request.POST.get('cropped_indices', '').split(',')
        cropped_indices = [int(i) for i in cropped_indices if i.isdigit()]
        
        try:
            # Update variant details
            variant.color = color.upper() if color else color  # Convert color to uppercase
            variant.price = price
            variant.discount_price = discount_price if discount_price else None
            variant.stock = 0  # Will be updated based on size stocks
            variant.save()
            
            existing_sizes = {size.name: size for size in variant.sizes.all()}
            total_stock = 0

            for size_code in new_sizes:
                stock_key = f'stock_{size_code}'
                stock_value = int(request.POST.get(stock_key, 0))
                
                if size_code in existing_sizes:
                    # Update existing size
                    size_obj = existing_sizes[size_code]
                    size_obj.stock = stock_value
                    size_obj.save()
                    del existing_sizes[size_code]
                else:
                    # Create new size
                    Size.objects.create(
                        product_variant=variant,
                        name=size_code,
                        stock=stock_value
                    )
                
                total_stock += stock_value

            # Handle sizes not in new_sizes
            for size_name, size_obj in existing_sizes.items():
                if size_obj.cartitem_set.exists():
                    # Set stock to 0 to retain cart items
                    size_obj.stock = 0
                    size_obj.save()
                else:
                    size_obj.delete()

            variant.stock = total_stock
            variant.save()
            
            # Handle new images
            if new_images:
                if len(new_images) < 4:
                    messages.error(request, 'Please upload at least 4 images.')
                    return redirect('editVariant', variant_id=variant.id)
                    
                # If new images are uploaded, delete old ones
                existing_images.delete()
                
                # Add new images - both original and cropped
                for index, image in enumerate(new_images):
                    # Check if this image was cropped
                    if index in cropped_indices:
                        # Use the cropped version if available
                        cropped_key = f'cropped_image_{index}'
                        if cropped_key in request.FILES:
                            image_to_save = request.FILES[cropped_key]
                        else:
                            image_to_save = image
                    else:
                        image_to_save = image
                    
                    ProductImage.objects.create(
                        variant=variant,
                        image=image_to_save,
                        is_primary=(index == 0)
                    )
        
            messages.success(request, 'Variant updated successfully.')
            return redirect('adminProducts')
            
        except Exception as e:
            messages.error(request, f'Error updating variant: {str(e)}')
            return redirect('editVariant', variant_id=variant.id)
    
    # Prepare size data for template
    size_data = []
    for size_code, size_name in Size.SIZE_CHIOCES:
        existing_size = existing_sizes.filter(name=size_code).first()
        size_data.append({
            'code': size_code,
            'name': size_name,
            'checked': existing_size is not None,
            'stock': existing_size.stock if existing_size else 0
        })
    
    context = {
        'product': product,
        'variant': variant,
        'size_data': size_data,
        'existing_images': existing_images
    }
    return render(request, 'editVariant.html', context)

@login_required(login_url='adminLogin')
def deleteProduct(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        try:
            # Set is_active to False instead of actually deleting
            product.is_active = False
            product.save()
            messages.success(request, 'Product blocked successfully')
        except Exception as e:
            messages.error(request, f'Error deleting product: {str(e)}')
        
    return redirect('adminProducts')

@login_required(login_url='adminLogin')
def deleteVariant(request, variant_id):
    if request.method == 'POST':
        variant = get_object_or_404(Variant, id=variant_id)
        try:
            # Set is_active to False instead of actually deleting
            variant.is_active = False
            variant.save()
            messages.success(request, 'Variant deleted successfully')
        except Exception as e:
            messages.error(request, f'Error deleting variant: {str(e)}')
    
    return redirect('adminProducts')

# @login_required(login_url='adminLogin')
# def toggle_product_status(request, product_id):
#     print(f"Toggling status for product ID: {product_id}")
#     try:
#         product = Product.objects.get(id=product_id)
#         product.is_active = not product.is_active
#         product.save()
        
       
#         return JsonResponse({'success': True, 'status': 'active' if product.is_active else 'blocked'})
#     except Product.DoesNotExist:
#         return JsonResponse({'success': False, 'error': 'Product not found'}, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
@login_required(login_url='adminLogin')
def toggleVariant(request, variant_id):
    if request.method == 'POST':
        variant = get_object_or_404(Variant, id=variant_id)
        try:
            variant.is_active = not variant.is_active
            variant.save()
            status = "unblocked" if variant.is_active else "blocked"
            messages.success(request, f'Variant successfully {status}')
        except Exception as e:
            messages.error(request, f'Error updating variant status: {str(e)}')
    return redirect('adminProducts')


############## orders   ####################

@login_required(login_url='adminLogin')
def admin_orders(request):
    # Handle initial page render (HTML)

    orders = Order.objects.select_related('user', 'payment_method').order_by('-created_at')

    search_query = request.GET.get('search', '')
    if search_query:
        orders = orders.filter(Q(user__username__icontains=search_query) | Q(order_id__icontains=search_query))
    
    paginator = Paginator(orders, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {'orders': page_obj, 'search_query': search_query}
    return render(request, 'adminOrders.html', context)


@login_required(login_url='adminLogin')
def admin_orders_api(request):
    try:
        # Fetch orders with related user, paymentmethod, and order items
        orders = Order.objects.select_related('user', 'payment_method').prefetch_related('order_items')
        
        # Apply search filter
        search_query = request.GET.get('search', '')
        if search_query:
            orders = orders.filter(Q(user__username__icontains=search_query) | Q(order_id__icontains=search_query))
        
        # Apply status filter
        status_filter = request.GET.get('status', 'all')
        if status_filter != 'all':
            orders = orders.filter(order_status__iexact=status_filter)
            
        # Apply payment method filter
        payment_filter = request.GET.get('payment_method', 'all')
        if payment_filter != 'all':
            orders = orders.filter(payment_method__name__iexact=payment_filter)
            
        # Apply sorting
        sort_field = request.GET.get('sort_field', 'created_at')
        sort_direction = request.GET.get('sort_direction', 'desc')
        
        # Map frontend sort fields to model fields if needed
        field_mapping = {
            'id': 'order_id',
            'username': 'user__username',
            'status': 'order_status',
            'subtotal': 'subtotal',
            'total_price': 'total_amount',
            'created_at': 'created_at'
        }
        
        db_sort_field = field_mapping.get(sort_field, 'created_at')
        if sort_direction == 'desc':
            db_sort_field = f'-{db_sort_field}'
            
        orders = orders.order_by(db_sort_field)
        
        # Paginate results
        per_page = int(request.GET.get('per_page', 10))  # Get per_page from request, default to 10
        paginator = Paginator(orders, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # Prepare data for JSON response
        orders_data = []
        for order in page_obj:
            order_items = [
                {
                    'product_name': item.product.name if item.product else 'N/A',
                    'variant_details': f"{item.variant.color}" if item.variant else 'N/A',
                    'quantity': item.quantity,
                    'price': float(item.price),
                    'status': item.status
                } for item in order.order_items.all()
            ]
            
            # Calculate total discount including both product and coupon discounts
            total_discount = float(order.product_discount) + float(order.coupon_discount)
            
            # Check if payment has expired and auto-cancel if needed
            from django.utils import timezone
            if (order.order_status == 'PENDING' and 
                order.payment_expiry_time and 
                timezone.now() > order.payment_expiry_time):
                # Auto-cancel the order if payment time expired
                order.order_status = 'CANCELLED'
                order.save(update_fields=['order_status'])
                # Create status history entry
                OrderStatusHistory.objects.get_or_create(
                    order=order,
                    old_status='PENDING',
                    new_status='CANCELLED',
                    defaults={
                        'reason': 'Auto-cancelled: Payment time window exceeded',
                        'changed_by': None
                    }
                )
            
            # Determine actual order status based on item statuses
            item_statuses = [item.status for item in order.order_items.all()]
            
            # If all items are returned, show RETURNED
            if item_statuses and all(status == 'RETURNED' for status in item_statuses):
                display_status = 'RETURNED'
            # If all items are cancelled, show CANCELLED
            elif item_statuses and all(status == 'CANCELLED' for status in item_statuses):
                display_status = 'CANCELLED'
            # If some items are returned or cancelled, show PARTIAL_RETURN or PARTIAL_CANCEL
            elif 'RETURNED' in item_statuses:
                display_status = 'PARTIAL_RETURN'
            elif 'CANCELLED' in item_statuses and not all(status == 'CANCELLED' for status in item_statuses):
                display_status = 'PARTIAL_CANCEL'
            else:
                display_status = order.order_status
            
            orders_data.append({
                'id': order.order_id,
                'username': order.user.username if order.user else 'N/A',
                'payment_method': order.payment_method.name if order.payment_method else 'N/A',
                'status': display_status,
                'subtotal': float(order.subtotal),
                'total_discount': total_discount,
                'total_price': float(order.total_amount) if order.total_amount else 0.0,
                'created_at': order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'items': order_items
            })
        
        return JsonResponse({
            'orders': orders_data,
            'total_pages': paginator.num_pages
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='adminLogin')
def order_detail(request, order_id):
    order = get_object_or_404(Order, order_id=order_id)
    
    # Determine actual order status based on item statuses
    item_statuses = [item.status for item in order.order_items.all()]
    
    # If all items are returned, the order is effectively returned
    if item_statuses and all(status == 'RETURNED' for status in item_statuses):
        actual_status = 'RETURNED'
    # If all items are cancelled, the order is effectively cancelled
    elif item_statuses and all(status == 'CANCELLED' for status in item_statuses):
        actual_status = 'CANCELLED'
    # If some items are returned
    elif 'RETURNED' in item_statuses:
        actual_status = 'PARTIAL_RETURN'
    # If some items are cancelled but not all
    elif 'CANCELLED' in item_statuses and not all(status == 'CANCELLED' for status in item_statuses):
        actual_status = 'PARTIAL_CANCEL'
    else:
        actual_status = order.order_status
    
    return render(request, 'order_detail.html', {
        'order': order,
        'actual_status': actual_status,
        'has_returned_items': 'RETURNED' in item_statuses,
        'has_cancelled_items': 'CANCELLED' in item_statuses
    })


logger = logging.getLogger(__name__)

@login_required(login_url='adminLogin')
@require_POST
def update_order_status(request, order_id):
    try:
        with transaction.atomic():
            data = json.loads(request.body)
            new_status = data.get('status')
            reason = data.get('reason')
            
            if not new_status:
                return JsonResponse({'success': False, 'error': 'Status is required'})
            
            order = get_object_or_404(Order, order_id=order_id)
            
            # Check if all items are cancelled
            active_items = order.order_items.exclude(status='CANCELLED').count()
            if active_items == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'Cannot modify status of fully cancelled orders'
                })
            
            # Remove CANCELLED from valid transitions as it's handled by user
            valid_status_transitions = {
                'PLACED': ['CONFIRMED'],
                'CONFIRMED': ['PROCESSING'],
                'PROCESSING': ['SHIPPED'],
                'SHIPPED': ['DELIVERED'],
                'DELIVERED': [],
                'CANCELLED': []
            }
            
            old_status = order.order_status
            if old_status in valid_status_transitions:
                if new_status not in valid_status_transitions[old_status]:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Invalid status transition: {old_status} -> {new_status}'
                    })
            
            # Create status history entry
            OrderStatusHistory.objects.create(
                order=order,
                old_status=old_status,
                new_status=new_status,
                reason=reason,
                changed_by=request.user
            )
            
            # Update order status
            order.order_status = new_status
            
            # If order is delivered and payment method is COD, mark payment as PAID
            if new_status == 'DELIVERED' and order.payment_method and order.payment_method.name.upper() == 'COD':
                order.payment_status = 'PAID'
            
            # Set appropriate timestamp based on status
            timestamp_field = {
                'CONFIRMED': 'confirmed_at',
                'PROCESSING': 'processed_at',
                'SHIPPED': 'shipped_at',
                'DELIVERED': 'delivered_at'
            }.get(new_status)
            
            if timestamp_field:
                setattr(order, timestamp_field, timezone.now())
            
            # Update only non-cancelled items status and timestamps
            active_items = order.order_items.exclude(status='CANCELLED')
            active_items.update(status=new_status)
            
            # Also update the timestamp for each active item
            if timestamp_field:
                current_time = timezone.now()
                for item in active_items:
                    setattr(item, timestamp_field, current_time)
                    item.save()
            
            order.save()
            
            # Send notification to customer
            send_order_status_notification(order)
            
            return JsonResponse({'success': True})
            
    except Exception as e:
        logger.error(f"Error updating order status: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='adminLogin')
@require_POST
def cancel_order(request, order_id):
    try:
        order = get_object_or_404(Order, order_id=order_id)
        if order.order_status not in ['DELIVERED', 'CANCELLED']:
            with transaction.atomic():
                order.order_status = 'CANCELLED'
                order.cancelled_at = timezone.now()
                order.save()
                # Update all order items status to CANCELLED
                order.order_items.all().update(status='CANCELLED', cancelled_at=timezone.now())
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'message': 'Order cannot be cancelled'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})    

@login_required(login_url='adminLogin')
@require_POST
def update_item_status(request):
    try:
        data = json.loads(request.body)
        item = get_object_or_404(OrderItem, id=data['item_id'])
        
        with transaction.atomic():
            item.status = data['status']
            
            # Set the appropriate timestamp based on status
            timestamp_fields = {
                'DELIVERED': 'delivered_at',
                'SHIPPED': 'shipped_at',
                'CANCELLED': 'cancelled_at',
                'PROCESSING': 'processed_at',
                'CONFIRMED': 'confirmed_at',
                'RETURNED': 'returned_at'
            }
            
            # Convert status to uppercase to match the keys in timestamp_fields
            status_upper = data['status'].upper()
            if status_upper in timestamp_fields:
                setattr(item, timestamp_fields[status_upper], timezone.now())
                
            if data['status'] in ['Cancelled', 'Returned'] and data.get('reason'):
                if data['status'] == 'Cancelled':
                    item.cancellation_reason = data['reason']
                else:
                    item.return_reason = data['reason']
            item.save()
            
            # Update order status if all items have the same status
            order = item.order
            all_items_status = set(order.order_items.values_list('status', flat=True))
            if len(all_items_status) == 1:
                order.order_status = data['status'].upper()
                order.save()
                
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='adminLogin')
def admin_coupons(request):
    """ Fetch all coupons ordered by newest first with usage statistics. """
    coupons = Coupon.objects.all().order_by('-created_at')
    
    # Add usage statistics to each coupon
    coupon_data = []
    for coupon in coupons:
        total_usage = CouponUsage.objects.filter(coupon=coupon).count()
        coupon_data.append({
            'coupon': coupon,
            'total_usage': total_usage
        })
    
    return render(request, 'admin_coupons.html', {'coupons': coupon_data})

@login_required(login_url='adminLogin')
@csrf_exempt
def add_coupon(request):
    """ Add a new coupon after validation. """
    if request.method == 'POST':
        try:
            code = request.POST.get('code', '').strip()
            coupon_type = request.POST.get('type')
            value = request.POST.get('value')
            minimum_purchase = request.POST.get('minimum_purchase')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            usage_limit = request.POST.get('usage_limit')
            is_active = request.POST.get('is_active') == 'on'

            # Validate coupon code (alphanumeric check)
            if not code.isalnum():
                messages.error(request, 'Coupon code must be alphanumeric.')
                return redirect('admin_coupons')

            # Check if coupon code already exists (case-sensitive)
            if Coupon.objects.filter(code=code).exists():
                messages.error(request, 'Coupon code already exists. Note: Coupon codes are case-sensitive.')
                return redirect('admin_coupons')

            # Create and save coupon
            Coupon.objects.create(
                code=code,
                type=coupon_type,
                value=value,
                minimum_purchase=minimum_purchase,
                start_date=start_date,
                end_date=end_date,
                usage_limit=usage_limit,
                is_active=is_active
            )
            messages.success(request, 'Coupon added successfully.')

        except Exception as e:
            messages.error(request, f'Error creating coupon: {str(e)}')

    return redirect('admin_coupons')

@login_required(login_url='adminLogin')
@require_POST
def delete_coupon(request, coupon_id):
    try:
        coupon = get_object_or_404(Coupon, id=coupon_id)
        coupon.delete()
        return JsonResponse({'success': True, 'message': 'Coupon deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
@login_required(login_url='adminLogin')
def apply_coupon(request):
    if request.method == 'POST':
        code = request.POST.get('coupon_code')
        cart = request.user.cart_set.filter(is_ordered=False).first()
        
        try:
            coupon = Coupon.objects.get(
                code__iexact=code,
                is_active=True,
                start_date__lte=timezone.now().date(),
                end_date__gte=timezone.now().date()
            )
            
            # Check minimum purchase
            if cart.total_price < coupon.minimum_purchase:
                return JsonResponse({
                    'success': False,
                    'message': f'Minimum purchase amount of ₹{coupon.minimum_purchase} required'
                })
            
            # Check usage limit
            usage_count = CouponUsage.objects.filter(
                coupon=coupon,
                user=request.user
            ).count()
            
            if usage_count >= coupon.usage_limit:
                return JsonResponse({
                    'success': False,
                    'message': 'Coupon usage limit exceeded'
                })
            
            # Calculate discount
            if coupon.type == 'percentage':
                discount = (cart.total_price * coupon.value) / 100
            else:
                discount = coupon.value
                
            return JsonResponse({
                'success': True,
                'discount': float(discount),
                'final_amount': float(cart.total_price - discount)
            })
            
        except Coupon.DoesNotExist:
            # Check if coupon exists but is expired or not active
            if Coupon.objects.filter(code__iexact=code).exists():
                coupon = Coupon.objects.get(code__iexact=code)
                if not coupon.is_active:
                    return JsonResponse({
                        'success': False,
                        'message': 'This coupon is no longer active'
                    })
                elif coupon.start_date > timezone.now().date():
                    return JsonResponse({
                        'success': False,
                        'message': f'This coupon is not valid until {coupon.start_date}'
                    })
                elif coupon.end_date < timezone.now().date():
                    return JsonResponse({
                        'success': False,
                        'message': 'This coupon has expired'
                    })
            return JsonResponse({
                'success': False,
                'message': 'Invalid coupon code'
            })

@login_required(login_url='adminLogin')
def get_coupon_details(request, coupon_id):
    try:
        coupon = get_object_or_404(Coupon, id=coupon_id)
        return JsonResponse({
            'code': coupon.code,
            'type': coupon.type,
            'value': str(coupon.value),
            'minimum_purchase': str(coupon.minimum_purchase),
            'start_date': coupon.start_date.strftime('%Y-%m-%d'),
            'end_date': coupon.end_date.strftime('%Y-%m-%d'),
            'usage_limit': coupon.usage_limit,
            'is_active': coupon.is_active
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='adminLogin')
def edit_coupon(request, coupon_id):
    if request.method == 'POST':
        try:
            coupon = get_object_or_404(Coupon, id=coupon_id)
            
            # Log old values for debugging
            old_usage_limit = coupon.usage_limit
            logger.info(f"Editing coupon {coupon.code} (ID: {coupon_id}). Old usage_limit: {old_usage_limit}")
            
            # Check if code exists for other coupons
            if Coupon.objects.filter(code=request.POST.get('code')).exclude(id=coupon_id).exists():
                messages.error(request, 'Coupon code already exists')
                return redirect('admin_coupons')

            coupon.code = request.POST.get('code')
            coupon.type = request.POST.get('type')
            coupon.value = request.POST.get('value')
            coupon.minimum_purchase = request.POST.get('minimum_purchase')
            coupon.start_date = request.POST.get('start_date')
            coupon.end_date = request.POST.get('end_date')
            coupon.usage_limit = request.POST.get('usage_limit')
            coupon.is_active = request.POST.get('is_active') == 'on'
            coupon.save()
            
            # Log new values
            logger.info(f"Coupon {coupon.code} updated. New usage_limit: {coupon.usage_limit}")
            
            messages.success(request, f'Coupon updated successfully. Usage limit changed from {old_usage_limit} to {coupon.usage_limit}')
        except Exception as e:
            logger.error(f"Error updating coupon {coupon_id}: {str(e)}", exc_info=True)
            messages.error(request, f'Error updating coupon: {str(e)}')
    return redirect('admin_coupons')

@login_required(login_url='adminLogin')
def coupon_usage_details(request, coupon_id):
    """Get detailed usage information for a specific coupon"""
    try:
        coupon = get_object_or_404(Coupon, id=coupon_id)
        usages = CouponUsage.objects.filter(coupon=coupon).select_related('user', 'order').order_by('-used_at')
        
        usage_data = []
        for usage in usages:
            usage_data.append({
                'username': usage.user.username,
                'email': usage.user.email,
                'order_id': usage.order.order_id,
                'used_at': usage.used_at.strftime('%d %b %Y, %I:%M %p')
            })
        
        return JsonResponse({
            'total_usage': usages.count(),
            'usage_limit': coupon.usage_limit,
            'unique_users': usages.values('user').distinct().count(),
            'usages': usage_data
        })
    except Exception as e:
        logger.error(f"Error fetching coupon usage details: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='adminLogin')
def admin_return_requests(request):
    # Just render the template - data will be loaded via API
    return render(request, 'return_requests.html')

@login_required(login_url='adminLogin')
def get_return_requests(request):
    """API endpoint to fetch return requests with pagination"""
    try:
        # Get query parameters
        page = int(request.GET.get('page', 1))
        search = request.GET.get('search', '')
        status_filter = request.GET.get('status', 'all')
        sort_field = request.GET.get('sort_field', 'created_at')
        sort_direction = request.GET.get('sort_direction', 'desc')
        items_per_page = int(request.GET.get('items_per_page', 10))
        
        # Start with all return requests
        query = ReturnRequest.objects.select_related('user', 'order_item', 'order_item__order', 'order_item__variant', 'order_item__variant__product')
        
        # Apply search filter if provided
        if search:
            query = query.filter(
                models.Q(order_item__order__order_id__icontains=search) |
                models.Q(user__username__icontains=search) |
                models.Q(user__email__icontains=search) |
                models.Q(order_item__variant__product__name__icontains=search)
            )
        
        # Apply status filter if provided
        if status_filter and status_filter.lower() != 'all':
            query = query.filter(status=status_filter.upper())
        
        # Apply sorting
        if sort_field == 'order_id':
            sort_key = 'order_item__order__order_id'
        elif sort_field == 'customer':
            sort_key = 'user__username'
        elif sort_field == 'product':
            sort_key = 'order_item__variant__product__name'
        elif sort_field == 'status':
            sort_key = 'status'
        else:
            sort_key = 'created_at'
            
        if sort_direction == 'desc':
            sort_key = f'-{sort_key}'
            
        query = query.order_by(sort_key)
        
        # Paginate the results
        paginator = Paginator(query, items_per_page)
        try:
            return_requests = paginator.page(page)
        except PageNotAnInteger:
            return_requests = paginator.page(1)
        except EmptyPage:
            return_requests = paginator.page(paginator.num_pages)
        
        # Format the data for the response
        return_requests_data = []
        for request in return_requests:
            # Safely get order_id
            order_id = 'N/A'
            if request.order_item:
                try:
                    order_id = request.order_item.order.order_id
                except (AttributeError, Exception):
                    order_id = 'N/A'
            
            # Safely get product information
            product_name = 'N/A'
            variant_color = ''
            if request.order_item:
                try:
                    product_name = request.order_item.variant.product.name if request.order_item.variant else request.order_item.product.name
                    variant_color = request.order_item.variant.color if request.order_item.variant else ''
                except (AttributeError, Exception):
                    product_name = 'N/A'
            
            # Calculate refund amount
            refund_amount = 0
            item_price = 0
            if request.order_item:
                try:
                    if request.order_item.order.payment_status == 'PAID':
                        refund_amount = float(request.order_item.get_refundable_amount())
                        item_price = float(request.order_item.price * request.order_item.quantity)
                except (AttributeError, Exception):
                    refund_amount = 0
                    item_price = 0
            
            return_requests_data.append({
                'id': request.id,
                'order_id': order_id,
                'customer': request.user.username if request.user else 'N/A',
                'customer_email': request.user.email if request.user else 'N/A',
                'product': f"{product_name} ({variant_color})" if variant_color else product_name,
                'reason': request.get_reason_display(),
                'description': request.description,
                'status': request.status,
                'created_at': request.created_at.strftime('%Y-%m-%d %H:%M'),
                'admin_response': request.admin_response or '',
                'refund_amount': refund_amount,
                'item_price': item_price,
                'quantity': request.order_item.quantity if request.order_item else 0
            })
        
        return JsonResponse({
            'success': True,
            'return_requests': return_requests_data,
            'total_items': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page
        })
    except Exception as e:
        error_message = f"Error fetching return requests: {str(e)}"
        logger.error(error_message, exc_info=True)
        
        # Add more detailed error information for debugging
        import traceback
        trace = traceback.format_exc()
        logger.error(f"Traceback: {trace}")
        
        return JsonResponse({
            'success': False,
            'message': error_message,
            'error_details': str(e)
        }, status=500)

@login_required(login_url='adminLogin')
def handle_return_request(request, return_request_id):
    try:
        print(f"Handling return request: {return_request_id}")
        print(f"Request method: {request.method}")
        print(f"Request path: {request.path}")
        
        data = json.loads(request.body)
        print(f"Request data: {data}")
        
        return_request = get_object_or_404(ReturnRequest, id=return_request_id)
        print(f"Found return request: {return_request.id}")
        
        action = data.get('action')
        admin_response = data.get('response', '')

        if action not in ['approve', 'reject']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid action specified'
            }, status=400)

        with transaction.atomic():
            if action == 'approve':

                # Update return request
                return_request.status = 'APPROVED'
                return_request.admin_response = admin_response
                return_request.save()
                
                # Update order item
                order_item = return_request.order_item
                order_item.status = 'RETURNED'  # Changed from 'Returned' to 'RETURNED'
                order_item.return_status = 'APPROVED'
                order_item.admin_response = admin_response
                order_item.returned_at = timezone.now()  # Set returned timestamp
                order_item.save()
                
                # Restore stock
                if order_item.size:
                    order_item.size.stock += order_item.quantity
                    order_item.size.save()
                
                # Recalculate order total based on remaining active items
                order = order_item.order
                active_items = order.order_items.exclude(status__in=['CANCELLED', 'RETURNED'])
                
                if active_items.exists():
                    # Calculate what customer paid for remaining items
                    # Sum up the actual paid amount for each remaining item
                    new_total = sum(item.get_refundable_amount() for item in active_items)
                    order.total_amount = new_total
                else:
                    # All items cancelled/returned
                    order.total_amount = Decimal('0')
                    order.order_status = 'RETURNED'
                
                order.save()
                    
                # Process refund if payment was made
                if order.payment_status == 'PAID':
                    wallet, created = Wallet.objects.get_or_create(user=order.user)
                    
                    # Calculate actual refundable amount (what customer actually paid for this item)
                    refund_amount = order_item.get_refundable_amount()
                    
                    wallet.balance += refund_amount
                    wallet.save()
                    
                    # Record transaction
                    item_price_before_coupon = order_item.price * order_item.quantity
                    
                    description = f'Refund for returned item #{order_item.id}'
                    if order.coupon_discount > 0:
                        description += f' (Item price: ₹{item_price_before_coupon:.2f}, Actual paid after all discounts: ₹{refund_amount:.2f})'
                    
                    WalletTransaction.objects.create(
                        wallet=wallet,
                        amount=refund_amount,
                        transaction_type='RETURN_REFUND',
                        description=description
                    )
            else:  # reject
                # Update return request
                return_request.status = 'REJECTED'
                return_request.admin_response = admin_response
                return_request.save()
                
                # Update order item
                order_item = return_request.order_item
                order_item.return_status = 'REJECTED'
                order_item.admin_response = admin_response
                order_item.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Return request {action}ed successfully'
            })
            
    except Exception as e:
        error_message = f"Error handling return request: {str(e)}"
        print(error_message)
        logger.error(error_message)
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'An error occurred: {str(e)}'
        }, status=400)

@login_required(login_url='adminLogin')
def get_return_request_info(request, return_request_id):
    """API endpoint to get information about a specific return request"""
    try:
        print(f"Getting info for return request: {return_request_id}")
        
        # Try to find the return request
        try:
            return_request = ReturnRequest.objects.select_related(
                'user', 'order_item', 'order_item__order', 'order_item__variant', 'order_item__variant__product'
            ).get(id=return_request_id)
        except ReturnRequest.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Return request with ID {return_request_id} not found'
            }, status=404)
        
        # Format the request data
        order_id = return_request.order_item.order.order_id if return_request.order_item and hasattr(return_request.order_item, 'order') and return_request.order_item.order else 'N/A'
        product_name = return_request.order_item.variant.product.name if return_request.order_item and hasattr(return_request.order_item, 'variant') and hasattr(return_request.order_item.variant, 'product') else 'N/A'
        variant_color = return_request.order_item.variant.color if return_request.order_item and hasattr(return_request.order_item, 'variant') else ''
        
        request_data = {
            'id': return_request.id,
            'order_id': order_id,
            'customer': return_request.user.username if return_request.user else 'N/A',
            'customer_email': return_request.user.email if return_request.user else 'N/A',
            'product': f"{product_name} ({variant_color})" if variant_color else product_name,
            'reason': return_request.get_reason_display(),
            'description': return_request.description,
            'status': return_request.status,
            'created_at': return_request.created_at.strftime('%Y-%m-%d %H:%M'),
            'admin_response': return_request.admin_response or ''
        }
        
        return JsonResponse({
            'success': True,
            'request': request_data
        })
        
    except Exception as e:
        error_message = f"Error getting return request info: {str(e)}"
        print(error_message)
        logger.error(error_message)
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'An error occurred: {str(e)}'
        }, status=500)


@login_required(login_url='adminLogin')
def admin_offers(request):
    today = timezone.now()
    offers = list(ProductOffer.objects.select_related('product').all()) + \
            list(CategoryOffer.objects.select_related('category').all())
    
    # Get all active products and categories for the dropdowns
    products = Product.objects.filter(is_active=True).order_by('name')
    categories = Category.objects.all().order_by('name')
    
    context = {
        'offers': offers,
        'products': products,
        'categories': categories,
        'today': today
    }
    return render(request, 'admin_offers.html', context)

@login_required(login_url='adminLogin')
def admin_product_offers(request):
    product_offers = ProductOffer.objects.select_related('product').all().order_by('-created_at')
    products = Product.objects.filter(is_active=True)
    
    context = {
        'product_offers': product_offers,
        'products': products
    }
    return render(request, 'admin_offers.html', context)

@login_required(login_url='adminLogin')
@require_POST
def add_product_offer(request):
    try:
        name = request.POST.get('name')
        product_id = request.POST.get('product')
        discount_percentage = request.POST.get('discount_percentage')
        start_date = timezone.make_aware(datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d'))
        end_date = timezone.make_aware(datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d'))
        is_active = request.POST.get('is_active') == 'on'
        
        # Validate input
        if not all([name, product_id, discount_percentage, start_date, end_date]):
            return JsonResponse({
                'success': False,
                'errors': {'form': 'All fields are required'}
            }, status=400)
        
        # Multiple offers are now allowed for the same product
        # The system will automatically apply the highest discount offer
        
        ProductOffer.objects.create(
            name=name,
            product_id=product_id,
            discount_percentage=discount_percentage,
            start_date=start_date,
            end_date=end_date,
            is_active=is_active
        )
        return JsonResponse({
            'success': True,
            'message': 'Product offer added successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required(login_url='adminLogin')
def get_product_offer(request, offer_id):
    try:
        offer = get_object_or_404(ProductOffer, id=offer_id)
        return JsonResponse({
            'name': offer.name,
            'product_id': offer.product.id,
            'discount_percentage': str(offer.discount_percentage),
            'start_date': offer.start_date.isoformat(),
            'end_date': offer.end_date.isoformat(),
            'is_active': offer.is_active
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='adminLogin')
@require_POST
def edit_product_offer(request, offer_id):
    try:
        offer = get_object_or_404(ProductOffer, id=offer_id)
        
        # Get form data
        name = request.POST.get('name')
        product_id = request.POST.get('product')
        discount_percentage = request.POST.get('discount_percentage')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Validate required fields
        if not all([name, product_id, discount_percentage, start_date, end_date]):
            return JsonResponse({
                'success': False,
                'errors': {'form': 'All fields are required'}
            }, status=400)
        
        try:
            # Parse and validate dates
            start_date = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end_date = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d'))
        except ValueError:
            return JsonResponse({
                'success': False,
                'errors': {'form': 'Invalid date format'}
            }, status=400)
        
        # Validate date order
        if start_date > end_date:
            return JsonResponse({
                'success': False,
                'errors': {'form': 'Start date must be before end date'}
            }, status=400)
            
        # Validate discount percentage
        try:
            discount_percentage = float(discount_percentage)
            if not (0 <= discount_percentage <= 100):
                return JsonResponse({
                    'success': False,
                    'errors': {'discount_percentage': 'Discount percentage must be between 0 and 100'}
                }, status=400)
        except ValueError:
            return JsonResponse({
                'success': False,
                'errors': {'discount_percentage': 'Invalid discount percentage'}
            }, status=400)
            
        # Update offer details
        offer.name = name
        offer.product_id = product_id
        offer.discount_percentage = discount_percentage
        offer.start_date = start_date
        offer.end_date = end_date
        offer.is_active = request.POST.get('is_active') == 'on'
    
        # Check for overlapping offers
        if ProductOffer.objects.filter(
            product_id=offer.product_id,
            start_date__lte=offer.end_date,
            end_date__gte=offer.start_date
        ).exclude(id=offer_id).exists():
            return JsonResponse({
                'success': False,
                'errors': {'form': 'An offer already exists for this product during the specified period'}
            }, status=400)
        
        offer.save()
        return JsonResponse({
            'success': True,
            'message': 'Product offer updated successfully'
        })
    except ProductOffer.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Product offer not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required(login_url='adminLogin')
@require_POST
def delete_product_offer(request, offer_id):
    try:
        offer = get_object_or_404(ProductOffer, id=offer_id)
        offer.delete()
        return JsonResponse({'success': True, 'message': 'Product offer deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required(login_url='adminLogin')
def admin_category_offers(request):
    offers = CategoryOffer.objects.select_related('category').all().order_by('-created_at')
    categories = Category.objects.filter(is_active=True)
    
    context = {
        'offers': offers,
        'categories': categories
    }
    return render(request, 'admin_offers.html', context)

@login_required(login_url='adminLogin')
@require_POST
def add_category_offer(request):
    try:
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        discount_percentage = request.POST.get('discount_percentage')
        start_date = timezone.make_aware(datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d'))
        end_date = timezone.make_aware(datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d'))
        is_active = request.POST.get('is_active') == 'on'
        
        # Validate input
        if not all([name, category_id, discount_percentage, start_date, end_date]):
            return JsonResponse({
                'success': False,
                'errors': {'form': 'All fields are required'}
            }, status=400)
        
        # Check for overlapping offers
        if CategoryOffer.objects.filter(
            category_id=category_id,
            start_date__lte=end_date,
            end_date__gte=start_date
        ).exists():
            return JsonResponse({
                'success': False,
                'errors': {'form': 'An offer already exists for this category during the specified period'}
            }, status=400)
        
        CategoryOffer.objects.create(
            name=name,
            category_id=category_id,
            discount_percentage=discount_percentage,
            start_date=start_date,
            end_date=end_date,
            is_active=is_active
        )
        return JsonResponse({
            'success': True,
            'message': 'Category offer added successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required(login_url='adminLogin')
def get_category_offer(request, offer_id):
    try:
        offer = get_object_or_404(CategoryOffer, id=offer_id)
        return JsonResponse({
            'name': offer.name,
            'category_id': offer.category.id,
            'discount_percentage': str(offer.discount_percentage),
            'start_date': offer.start_date.isoformat(),
            'end_date': offer.end_date.isoformat(),
            'is_active': offer.is_active
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='adminLogin')
@require_POST
def edit_category_offer(request, offer_id):
    try:
        offer = get_object_or_404(CategoryOffer, id=offer_id)
        
        # Get and validate required fields
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        discount_percentage = request.POST.get('discount_percentage')
        start_date = timezone.make_aware(datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d'))
        end_date = timezone.make_aware(datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d'))
        
        # Validate required fields
        if not all([name, category_id, discount_percentage, start_date, end_date]):
            return JsonResponse({
                'success': False,
                'errors': {'form': 'All fields are required'}
            }, status=400)
            
        # Update offer details
        offer.name = name
        offer.category_id = category_id
        offer.discount_percentage = discount_percentage
        offer.start_date = start_date
        offer.end_date = end_date
        offer.is_active = request.POST.get('is_active') == 'on'
        
        # Check for overlapping offers
        if CategoryOffer.objects.filter(
            category_id=offer.category_id,
            start_date__lte=offer.end_date,
            end_date__gte=offer.start_date
        ).exclude(id=offer_id).exists():
            return JsonResponse({
                'success': False,
                'errors': {'form': 'An offer already exists for this category during the specified period'}
            }, status=400)
        
        offer.save()
        return JsonResponse({
            'success': True,
            'message': 'Category offer updated successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required(login_url='adminLogin')
@require_POST
def delete_category_offer(request, offer_id):
    try:
        offer = get_object_or_404(CategoryOffer, id=offer_id)
        offer.delete()
        return JsonResponse({'success': True, 'message': 'Category offer deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required(login_url='adminLogin')
def get_active_products(request):
    try:
        products = Product.objects.filter(is_active=True).values('id', 'name')
        return JsonResponse(list(products), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required(login_url='adminLogin')
def toggle_category_status(request, category_id):
    print(f"Toggling category {category_id}")
    if request.method == 'POST':
        category = get_object_or_404(Category, id=category_id)
        category.is_active = not category.is_active
        category.save()
        print(f"New status: {'active' if category.is_active else 'blocked'}")
        return JsonResponse({
            'success': True,
            'status': 'active' if category.is_active else 'blocked'
        })
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

@login_required(login_url='adminLogin')
def sales_report(request):
    # Get report type and date range from request
    report_type = request.GET.get('report_type', 'daily')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    download_format = request.GET.get('download')

    # Initialize the date range based on report type
    today = timezone.now()
    if report_type == 'daily':
        start_date = today.date()
        end_date = start_date
    elif report_type == 'weekly':
        start_date = today.date() - timezone.timedelta(days=today.weekday())
        end_date = start_date + timezone.timedelta(days=6)
    elif report_type == 'monthly':
        start_date = today.date().replace(day=1)
        end_date = (start_date + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(days=1)
    elif report_type == 'custom' and start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start_date = today.date()
        end_date = start_date

    # Filter orders based on date range and delivery status
    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        order_status='DELIVERED'
    ).order_by('-created_at')

    # Calculate summary statistics
    total_orders = orders.count()
    total_sales = sum(order.total_amount for order in orders)
    total_discounts = sum(order.total_discount for order in orders)

    # Handle report downloads
    if download_format:
        if download_format == 'pdf':
            # Generate PDF report with professional formatting
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="helmax_sales_report_{start_date}_{end_date}.pdf"'
            
            # Use our professional PDF generator function
            return generate_sales_report_pdf(
                response=response,
                orders=orders,
                total_orders=total_orders,
                total_sales=total_sales,
                total_discounts=total_discounts,
                start_date=start_date,
                end_date=end_date
            )
            

            
        elif download_format == 'excel':
            # Generate Excel report
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.xlsx"'
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Sales Report'
            
            # Add headers
            headers = ['Order ID', 'Date', 'Customer', 'Items', 'Subtotal', 'Discount', 'Total', 'Status']
            worksheet.append(headers)
            
            # Add data
            for order in orders:
                worksheet.append([
                    order.order_id,
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.user.username,
                    order.order_items.count(),
                    float(order.total_amount),
                    float(order.total_discount),
                    float(order.total_amount),
                    order.order_status
                ])
            
            # Style the worksheet
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            workbook.save(response)
            return response

    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_sales': total_sales,
        'total_discounts': total_discounts,
        'report_type': report_type,
        'start_date': start_date,
        'end_date': end_date
    }
    
    return render(request, 'sales_report.html', context)

@login_required(login_url='adminLogin')
def admin_brand(request):
    # Fetch all brands ordered by ID in descending order
    brands = Brand.objects.all().order_by("-id")
    
    # Search functionality for brands
    search_query = request.GET.get('search', '')
    if search_query:
        
        brands = brands.filter(name__icontains=search_query)
    
    # Paginate brands
    paginator = Paginator(brands, 10)  # Show 10 brands per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Prepare context
    context = {
        'brands': page_obj,  # Pass the paginated brands
        'search_query': search_query,  # Pass the search query
    }
    
    # Render the template with the context
    return render(request, 'adminBrand.html', context)

@login_required(login_url='adminLogin')
def add_brand(request):
    if request.method == "POST":
        name = request.POST.get('brand_name')
        if name and not Brand.objects.filter(name__iexact=name).exists():
            Brand.objects.create(name=name)
        return redirect('admin_brand')
    return redirect('admin_brand')

@login_required(login_url='adminLogin')
def edit_brand(request, brand_id):
    brand = get_object_or_404(Brand, id=brand_id)
    if request.method == "POST":
        
        name = request.POST.get('brand_name')
        if name and not Brand.objects.filter(name__iexact=name).exclude(id=brand_id).exists():
            brand.name = name
            brand.save()
        return redirect('admin_brand')
    return render(request, 'editBrand.html', {'brand': brand})

@login_required(login_url='adminLogin')
def toggle_brand_status(request, brand_id):
    if request.method == 'POST':
        brand = get_object_or_404(Brand, id=brand_id)
        brand.is_active = not brand.is_active
        brand.save()
        return JsonResponse({
            'success': True,
            'status': 'active' if brand.is_active else 'blocked'
        })
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

@login_required(login_url='adminLogin')
def adminProducts(request):
    # Fetch all products ordered by ID
    products = Product.objects.all().order_by('id')
    
    # Pagination Logic
    paginator = Paginator(products, 10)  # Show 10 products per page
    page_number = request.GET.get('page')  # Get the current page number from the request
    page_obj = paginator.get_page(page_number)  # Get the page object

    context = {
        'page_obj': page_obj,  # Pass the paginated page object to the template
    }
    return render(request, 'adminProducts.html', context)


@login_required(login_url='adminLogin')
def addProducts(request):
    categories = Category.objects.filter(is_active=True)
    brands = Brand.objects.filter(is_active=True)
    
    if request.method == 'POST':
        # Get and sanitize input
        name = request.POST.get('name', '').strip()  # Remove leading/trailing whitespace
        brand_id = request.POST.get('brand')
        category_id = request.POST.get('category')
        description = request.POST.get('description', '').strip()

        # Validate all fields are not empty
        if not name or not brand_id or not category_id or not description:
            messages.error(request, "All fields are required.")
            return render(request, 'addProducts.html', {'categories': categories, 'brands': brands})

        # # Validate product name (no symbols or only spaces)
        # if not name.replace(" ", "").isalnum():
        #     messages.error(request, "Product name cannot contain only symbols or spaces.")
        #     return render(request, 'addProducts.html', {'categories': categories, 'brands': brands})

        # Case-insensitive uniqueness check
        if Product.objects.filter(name__iexact=name).exists():
            messages.error(request, f"A product with the name '{name}' already exists (case-insensitive).")
            return render(request, 'addProducts.html', {'categories': categories, 'brands': brands})

        # Create product if valid
        try:
            category = get_object_or_404(Category, id=category_id)
            brand = get_object_or_404(Brand, id=brand_id)
            
            Product.objects.create(
                name=name,
                brand=brand,
                category=category,
                description=description
            )
            messages.success(request, "Product added successfully!")
            return redirect('adminProducts')
            
        except Exception as e:
            messages.error(request, f"Error creating product: {str(e)}")

    return render(request, 'addProducts.html', {'categories': categories, 'brands': brands})

@login_required(login_url='adminLogin')
def addVariant(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        color = request.POST.get('color')
        price = request.POST.get('price')
        discount_price = request.POST.get('discount_price')
        sizes = request.POST.getlist('sizes')
        images = request.FILES.getlist('images')
        primary_image_index = int(request.POST.get('primary_image_index', 0))  # New field

        try:
            with transaction.atomic():  # Use transaction to ensure data consistency
                # Create variant
                variant = Variant.objects.create(
                    product=product,
                    color=color,
                    price=price,
                    discount_price=discount_price if discount_price else None
                )

                # Create sizes with their respective stocks
                for size in sizes:
                    stock_for_size = request.POST.get(f'stock_{size}', 0)
                    stock_value = int(stock_for_size) if stock_for_size else 0
                    
                    Size.objects.create(
                        product_variant=variant,
                        name=size,
                        stock=stock_value
                    )

                # Handle images
                for index, image in enumerate(images):
                    ProductImage.objects.create(
                        variant=variant,
                        image=image,
                        is_primary=(index == primary_image_index)  # Set primary based on selected index
                    )

            messages.success(request, 'Variant added successfully.')
        except Exception as e:
            messages.error(request, f'Error adding variant: {str(e)}')

        return redirect('adminProducts')

    context = {
        'product': product,
        'size_choices': Size.SIZE_CHIOCES
    }
    return render(request, 'addVariant.html', context)

@login_required(login_url='adminLogin')
def editProduct(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    categories = Category.objects.filter(is_active=True)
    brands = Brand.objects.filter(is_active=True)
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            brand_id = request.POST.get('brand')
            category_id = request.POST.get('category')
            description = request.POST.get('description')

            product.name = name
            product.brand = get_object_or_404(Brand, id=brand_id)
            product.category = get_object_or_404(Category, id=category_id)
            product.description = description
            product.save()

            return redirect('adminProducts')
        except Exception as e:
            print(f"Error: {e}")
    
    context = {
        'product': product,
        'categories': categories,
        'brands': brands
    }
    return render(request, 'editProduct.html', context)

@login_required(login_url='adminLogin')
def editVariant(request, variant_id):
    variant = get_object_or_404(Variant, id=variant_id)
    product = variant.product
    existing_sizes = variant.sizes.all()
    existing_images = variant.images.all()
    
    if request.method == 'POST':
        color = request.POST.get('color')
        price = request.POST.get('price')
        discount_price = request.POST.get('discount_price')
        new_sizes = request.POST.getlist('sizes')
        new_images = request.FILES.getlist('images')
        
        try:
            # Update variant details
            variant.color = color.upper() if color else color  # Convert color to uppercase
            variant.price = price
            variant.discount_price = discount_price if discount_price else None
            variant.stock = 0  # Will be updated based on size stocks
            variant.save()
            
            existing_sizes = {size.name: size for size in variant.sizes.all()}
            total_stock = 0

            for size_code in new_sizes:
                stock_key = f'stock_{size_code}'
                stock_value = int(request.POST.get(stock_key, 0))
                
                if size_code in existing_sizes:
                    # Update existing size
                    size_obj = existing_sizes[size_code]
                    size_obj.stock = stock_value
                    size_obj.save()
                    del existing_sizes[size_code]
                else:
                    # Create new size
                    Size.objects.create(
                        product_variant=variant,
                        name=size_code,
                        stock=stock_value
                    )
                
                total_stock += stock_value

            # Handle sizes not in new_sizes
            for size_name, size_obj in existing_sizes.items():
                if size_obj.cartitem_set.exists():
                    # Set stock to 0 to retain cart items
                    size_obj.stock = 0
                    size_obj.save()
                else:
                    size_obj.delete()

            variant.stock = total_stock
            variant.save()
            
            # Handle new images
            if new_images:
                if len(new_images) < 4:
                    messages.error(request, 'Please upload at least 4 images.')
                    return redirect('editVariant', variant_id=variant.id)
                    
                # If new images are uploaded, delete old ones
                existing_images.delete()
                
                # Add new images
                for index, image in enumerate(new_images):
                    ProductImage.objects.create(
                        variant=variant,
                        image=image,
                        is_primary=(index == 0)
                    )
        
            messages.success(request, 'Variant updated successfully.')
            return redirect('adminProducts')
            
        except Exception as e:
            messages.error(request, f'Error updating variant: {str(e)}')
            return redirect('editVariant', variant_id=variant.id)
    
    # Prepare size data for template
    size_data = []
    for size_code, size_name in Size.SIZE_CHIOCES:
        existing_size = existing_sizes.filter(name=size_code).first()
        size_data.append({
            'code': size_code,
            'name': size_name,
            'checked': existing_size is not None,
            'stock': existing_size.stock if existing_size else 0
        })
    
    context = {
        'product': product,
        'variant': variant,
        'size_data': size_data,
        'existing_images': existing_images
    }
    return render(request, 'editVariant.html', context)
@login_required(login_url='adminLogin')
def deleteProduct(request, product_id):
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id)
        try:
            # Set is_active to False instead of actually deleting
            product.is_active = False
            product.save()
            messages.success(request, 'Product blocked successfully')
        except Exception as e:
            messages.error(request, f'Error deleting product: {str(e)}')
        
    return redirect('adminProducts')

@login_required(login_url='adminLogin')
def deleteVariant(request, variant_id):
    if request.method == 'POST':
        variant = get_object_or_404(Variant, id=variant_id)
        try:
            # Set is_active to False instead of actually deleting
            variant.is_active = False
            variant.save()
            messages.success(request, 'Variant deleted successfully')
        except Exception as e:
            messages.error(request, f'Error deleting variant: {str(e)}')
    
    return redirect('adminProducts')


    
@login_required(login_url='adminLogin')
def toggleVariant(request, variant_id):
    if request.method == 'POST':
        variant = get_object_or_404(Variant, id=variant_id)
        try:
            variant.is_active = not variant.is_active
            variant.save()
            status = "unblocked" if variant.is_active else "blocked"
            messages.success(request, f'Variant successfully {status}')
        except Exception as e:
            messages.error(request, f'Error updating variant status: {str(e)}')
    return redirect('adminProducts')


@login_required(login_url='adminLogin')
def admin_wallet(request):
    """View to display all wallet transactions for admin"""
    try:
        # Get search and filter parameters
        search_query = request.GET.get('search', '').strip()
        type_filter = request.GET.get('type', '').strip()
        
        # Start with all transactions with related wallets and users
        transactions = WalletTransaction.objects.all()\
            .select_related('wallet', 'wallet__user', 'order')
        
        # Apply search filter if provided
        if search_query:
            transactions = transactions.filter(
                models.Q(id__icontains=search_query) |
                models.Q(wallet__user__username__icontains=search_query) |
                models.Q(wallet__user__email__icontains=search_query) |
                models.Q(description__icontains=search_query)
            )
        
        # Apply type filter if provided
        if type_filter:
            transactions = transactions.filter(transaction_type=type_filter)
        
        # Order by most recent first
        transactions = transactions.order_by('-created_at')
        
        # Pagination
        paginator = Paginator(transactions, 20)  # Show 20 transactions per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Pre-select the type filter if it was provided
        context = {
            'transactions': page_obj,
            'search_query': search_query,
            'type_filter': type_filter
        }
        
        return render(request, 'admin_wallet.html', context)
        
    except Exception as e:
        logger.error(f"Error in admin wallet view: {str(e)}")
        messages.error(request, "An error occurred while loading wallet transactions.")
        return redirect('customers')


@login_required(login_url='adminLogin')
def admin_wallet_transaction_detail(request, transaction_id):
    """View to display detailed information about a specific wallet transaction"""
    try:
        # Get transaction with related wallet, user and order
        transaction = get_object_or_404(
            WalletTransaction.objects.select_related('wallet', 'wallet__user', 'order'),
            id=transaction_id
        )
        
        context = {
            'transaction': transaction,
        }
        
        return render(request, 'admin_wallet_transaction_detail.html', context)
        
    except Exception as e:
        logger.error(f"Error in transaction detail view: {str(e)}")
        messages.error(request, "An error occurred while loading transaction details.")
        return redirect('admin_wallet')



from manager.models import Order, OrderItem, Product, Category, Brand
@login_required(login_url='adminLogin')
def admin_dashboard(request):
    # Get pending returns count for the sidebar notification
    pending_returns_count = ReturnRequest.objects.filter(status='PENDING').count()
    
    # Get pending orders count
    pending_orders_count = Order.objects.filter(order_status='PENDING').count()
    
    return render(request, 'admin_dashboard.html', {
        'pending_returns_count': pending_returns_count,
        'pending_orders_count': pending_orders_count
    })
@login_required(login_url='adminLogin')
def dashboard_data(request):
    try:
        filter_type = request.GET.get('filter_type', 'monthly')
        today = timezone.now()
        
        # Set time range based on filter
        if filter_type == 'yearly':
            start_date = today - timedelta(days=365)
            date_format = '%Y-%m'
        elif filter_type == 'monthly':
            start_date = today - timedelta(days=30)
            date_format = '%Y-%m-%d'
        elif filter_type == 'weekly':
            start_date = today - timedelta(days=7)
            date_format = '%Y-%m-%d'
        else:  # daily
            start_date = today - timedelta(days=1)
            date_format = '%Y-%m-%d %H:00'
        
        # Get metrics data
        total_revenue = Order.objects.filter(
            created_at__gte=start_date,
            order_status='DELIVERED'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        total_orders = Order.objects.filter(
            created_at__gte=start_date
        ).count()
        
        new_customers = User.objects.filter(
            date_joined__gte=start_date
        ).count()
        
        # Get sales data
        sales_data = Order.objects.filter(
            created_at__gte=start_date,
            order_status='DELIVERED'
        ).extra(
            select={'date': f"to_char(created_at, '{date_format}')"}
        ).values('date').annotate(
            total=Sum('total_amount')
        ).order_by('date')
        
        # Get top products
        top_products = OrderItem.objects.filter(
            order__order_status='DELIVERED',
            order__created_at__gte=start_date
        ).values(
            'variant__product__name'
        ).annotate(
            units_sold=Sum('quantity'),
            revenue=Sum('total_price')
        ).order_by('-revenue')[:5]
        
        # Get top categories
        top_categories = OrderItem.objects.filter(
            order__order_status='DELIVERED',
            order__created_at__gte=start_date
        ).values(
            'variant__product__category__name'
        ).annotate(
            units_sold=Sum('quantity'),
            revenue=Sum('total_price')
        ).order_by('-revenue')[:5]
        
        # Get top brands
        top_brands = OrderItem.objects.filter(
            order__order_status='DELIVERED',
            order__created_at__gte=start_date
        ).values(
            'variant__product__brand__name'
        ).annotate(
            units_sold=Sum('quantity'),
            revenue=Sum('total_price')
        ).order_by('-revenue')[:5]
        
        # Prepare the response data
        response_data = {
            'metrics': {
                'total_revenue': float(total_revenue),
                'total_orders': total_orders,
                'new_customers': new_customers
            },
            'sales_data': {
                'labels': [item['date'] for item in sales_data],
                'values': [float(item['total']) for item in sales_data]
            },
            'top_products': {
                'products': [{
                    'name': item['variant__product__name'],
                    'units_sold': item['units_sold'],
                    'revenue': float(item['revenue'])
                } for item in top_products]
            },
            'top_categories': {
                'categories': [{
                    'name': item['variant__product__category__name'],
                    'units_sold': item['units_sold'],
                    'revenue': float(item['revenue'])
                } for item in top_categories]
            },
            'top_brands': {
                'labels': [item['variant__product__brand__name'] for item in top_brands],
                'values': [float(item['revenue']) for item in top_brands],
                'brands': [{
                    'name': item['variant__product__brand__name'],
                    'units_sold': item['units_sold'],
                    'revenue': float(item['revenue'])
                } for item in top_brands]
            }
        }
        return JsonResponse(response_data)
    except Exception as e:
        logging.error(f"Error in dashboard_data: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)
@login_required(login_url='adminLogin')
def generate_ledger(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Convert string dates to datetime objects
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    
    # Get all orders within date range
    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).order_by('created_at')
    
    # Calculate summary statistics
    total_sales = orders.filter(order_status='DELIVERED').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_orders = orders.count()
    delivered_orders = orders.filter(order_status='DELIVERED').count()
    cancelled_orders = orders.filter(order_status='CANCELLED').count()
    
    # Prepare ledger entries
    ledger_entries = []
    running_balance = 0
    
    for order in orders:
        entry = {
            'date': order.created_at.strftime('%Y-%m-%d'),
            'order_id': order.order_id,
            'description': f"Order {order.order_id} - {order.order_status}",
            'debit': 0,
            'credit': 0,
            'balance': 0
        }
        
        # Add to credit if order is delivered (income)
        if order.order_status == 'DELIVERED':
            entry['credit'] = float(order.total_amount)
            running_balance += float(order.total_amount)
        # Add to debit if order is cancelled or returned (expense/refund)
        elif order.order_status in ['CANCELLED', 'RETURNED']:
            entry['debit'] = float(order.total_amount)
            running_balance -= float(order.total_amount)
        
        entry['balance'] = running_balance
        ledger_entries.append(entry)
    
    context = {
        'ledger_entries': ledger_entries,
        'total_sales': total_sales,
        'total_orders': total_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        'start_date': start_date,
        'end_date': end_date
    }
    
    return render(request, 'admin_ledger.html', context)